"""
core/generador_ir17_reporte.py — Reporte Visual Ejecutivo del IR-17
==================================================================
Genera un archivo Excel formateado con las retenciones mensuales,
incluyendo gráficos de tendencia y formato condicional profesional.
"""
import sys, os

# Resolución dinámica de ruta del proyecto
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

import pandas as pd
from database import SessionLocal, Empresa, DgiiIr17
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")

def generar_reporte_visual_ir17(rnc: str, anio: int):
    print(f"\n[Visual] Generando Reporte Excel IR-17 para RNC {rnc}...")
    
    db = SessionLocal()
    try:
        empresa = db.query(Empresa).filter_by(rnc=rnc).first()
        if not empresa: return None
        
        ir17_data = db.query(DgiiIr17).filter(
            DgiiIr17.empresa_id == empresa.id,
            DgiiIr17.periodo.like(f"{anio}%")
        ).order_by(DgiiIr17.periodo).all()
        
        if not ir17_data:
            print("  [!] No hay datos de IR-17 para reporte.")
            return None
            
        # 1. Preparar DataFrame
        data = []
        meses_map = {
            "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
            "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
            "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
        }
        
        for reg in ir17_data:
            mes_code = reg.periodo[4:]
            data.append({
                "Mes": meses_map.get(mes_code, mes_code),
                "Alquileres (10%)": float(reg.alquileres),
                "Honorarios (10%)": float(reg.honorarios),
                "Servicios Técnicos (2%)": float(reg.servicios_tecnicos),
                "Otros Pagos": float(reg.otros_pagos),
                "ITBIS Retenido": float(reg.itbis_retenido_terceros),
                "TOTAL A PAGAR": float(reg.total_a_pagar)
            })
            
        df = pd.DataFrame(data)
        
        # 2. Crear Excel con openpyxl para control total de diseño
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        filename = f"REPORTE_IR17_{anio}_{rnc}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        writer = pd.ExcelWriter(filepath, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Detalle Mensual')
        
        workbook = writer.book
        sheet = workbook['Detalle Mensual']
        
        # 3. Diseño Premium
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # Estilizar cabecera
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Formato de moneda y ajustes de columna
        for row in sheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=7):
            for cell in row:
                cell.number_format = '#,##0.00'
                
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column_letter].width = max_length + 2

        # 4. Añadir Gráfico de Barras (Tendencia)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Tendencia Retenciones Mensuales {anio}"
        chart.y_axis.title = 'Monto (RD$)'
        chart.x_axis.title = 'Mes'
        
        # Data para el gráfico (excluyendo el total para no sesgar)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(df)+1)
        data_chart = Reference(sheet, min_col=2, max_col=6, min_row=1, max_row=len(df)+1)
        
        chart.add_data(data_chart, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 10
        
        # Colocar el gráfico a la derecha de la tabla
        sheet.add_chart(chart, "I2")
        
        writer.close()
        print(f"  [OK] Reporte Generado: {filepath}")
        return filepath
        
    finally:
        db.close()

if __name__ == "__main__":
    generar_reporte_visual_ir17("130826552", 2025)
